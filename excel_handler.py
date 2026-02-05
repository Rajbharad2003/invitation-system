"""
Excel Handler Module
Handles reading contact information from Excel files
"""
import openpyxl
from typing import List, Dict
import os


def read_contacts(excel_path: str) -> List[Dict[str, str]]:
    """
    Read contacts from an Excel file.
    
    Expected format:
    - Column A (or first column): Name
    - Column B (or second column): Mobile number
    
    Args:
        excel_path: Path to the Excel file (.xlsx)
    
    Returns:
        List of dictionaries with 'name' and 'mobile' keys
    """
    contacts = []
    
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_path)
    
    # Get the active sheet
    sheet = workbook.active
    
    # Find the header row and column indices
    name_col = None
    mobile_col = None
    header_row = 1
    
    # Check first row for headers
    for col in range(1, sheet.max_column + 1):
        cell_value = str(sheet.cell(row=1, column=col).value or "").lower().strip()
        if cell_value in ['name', 'names', 'नाव', 'नाम']:
            name_col = col
        elif cell_value in ['mobile', 'phone', 'number', 'mobile number', 'phone number', 'मोबाइल', 'फोन']:
            mobile_col = col
    
    # If headers not found, assume first two columns
    if name_col is None:
        name_col = 1
        header_row = 0  # No header, start from row 1
    if mobile_col is None:
        mobile_col = 2
    
    # Read data rows
    start_row = header_row + 1 if header_row > 0 else 1
    
    for row in range(start_row, sheet.max_row + 1):
        name_cell = sheet.cell(row=row, column=name_col).value
        mobile_cell = sheet.cell(row=row, column=mobile_col).value
        
        # Skip empty rows
        if not name_cell or not mobile_cell:
            continue
        
        # Clean up mobile number
        # Handle numbers stored as floats in Excel (e.g., 9904276722.0)
        mobile = str(mobile_cell).strip()
        if '.' in mobile:
            mobile = mobile.split('.')[0]  # Remove decimal part
        
        # Remove any spaces, dashes, or other non-digit characters
        mobile = ''.join(c for c in mobile if c.isdigit())
        
        # Keep only the raw 10-digit number (don't add country code here)
        # Country code will be added in whatsapp_handler
        if len(mobile) > 10 and mobile.startswith('91'):
            mobile = mobile[2:]  # Remove 91 prefix if present
        elif len(mobile) > 10:
            mobile = mobile[-10:]  # Take last 10 digits
        
        contacts.append({
            'name': str(name_cell).strip(),
            'mobile': mobile
        })
    
    workbook.close()
    
    return contacts


def validate_excel_file(excel_path: str) -> Dict:
    """
    Validate an Excel file for the expected format.
    
    Args:
        excel_path: Path to the Excel file
    
    Returns:
        Dictionary with 'valid' boolean and 'message' string
    """
    if not os.path.exists(excel_path):
        return {'valid': False, 'message': 'File does not exist'}
    
    if not excel_path.endswith('.xlsx'):
        return {'valid': False, 'message': 'File must be an .xlsx file'}
    
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        if sheet.max_row < 2:
            return {'valid': False, 'message': 'Excel file appears to be empty or has only headers'}
        
        if sheet.max_column < 2:
            return {'valid': False, 'message': 'Excel file needs at least 2 columns (Name and Mobile)'}
        
        workbook.close()
        
        # Try reading contacts
        contacts = read_contacts(excel_path)
        
        if len(contacts) == 0:
            return {'valid': False, 'message': 'No valid contacts found in the Excel file'}
        
        return {
            'valid': True, 
            'message': f'Found {len(contacts)} contacts',
            'count': len(contacts)
        }
        
    except Exception as e:
        return {'valid': False, 'message': f'Error reading Excel file: {str(e)}'}
